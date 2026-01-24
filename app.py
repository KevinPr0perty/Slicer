import io
import math
import time
import zipfile

import streamlit as st
from openpyxl import load_workbook


def split_using_template_keep_format(
    template_bytes: bytes,
    chunk_size: int = 999,
    header_rows: int = 2,
    sheet_name: str | None = None,
    progress_callback=None,
):
    """
    Preserves formatting by reloading the original template for each part,
    trimming rows, then writing ONLY values into the formatted template rows.
    """
    # Load once to count data rows & figure out column count / sheet
    wb0 = load_workbook(io.BytesIO(template_bytes))
    ws0 = wb0[sheet_name] if sheet_name else wb0.active

    max_row = ws0.max_row
    max_col = ws0.max_column

    data_start = header_rows + 1  # row 3
    data_rows = max(0, max_row - header_rows)
    if data_rows <= 0:
        raise ValueError("No data rows found (expected data starting at row 3).")

    parts = math.ceil(data_rows / chunk_size)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        start_time = time.time()

        for part_idx in range(parts):
            # Reload template fresh each time to preserve ALL workbook formatting
            wb = load_workbook(io.BytesIO(template_bytes))
            ws = wb[sheet_name] if sheet_name else wb.active

            start_idx = part_idx * chunk_size
            end_idx = min((part_idx + 1) * chunk_size, data_rows)
            this_len = end_idx - start_idx

            # Keep total rows <= header_rows + this_len
            # Delete all rows after the last needed data row (preserves top formatting)
            keep_through = header_rows + this_len
            if ws.max_row > keep_through:
                ws.delete_rows(keep_through + 1, ws.max_row - keep_through)

            # Write ONLY values into existing formatted cells
            # (No style copying, so formatting stays identical)
            out_r = data_start
            for i, k in enumerate(range(start_idx, end_idx), start=1):
                src_r = data_start + k
                for c in range(1, max_col + 1):
                    ws.cell(row=out_r, column=c).value = ws0.cell(row=src_r, column=c).value
                out_r += 1

                # Progress update (throttle)
                if progress_callback and (i == 1 or i == this_len or i % 25 == 0):
                    overall = (part_idx + i / max(this_len, 1)) / parts
                    progress_callback(overall, part_idx + 1, parts, i, this_len, time.time() - start_time)

            # Save part to bytes, add to zip
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            zf.writestr(f"part_{part_idx + 1}_of_{parts}.xlsx", bio.read())

            if progress_callback:
                progress_callback((part_idx + 1) / parts, part_idx + 1, parts, this_len, this_len, time.time() - start_time)

    zip_buffer.seek(0)
    return zip_buffer.read(), parts


# ---------- Streamlit UI ----------
st.set_page_config(page_title="Temu Excel Splitter (Local)", layout="centered")
st.title("Temu Excel Splitter (Keeps Template Formatting)")

st.write("This version preserves the exact Temu template formatting by editing the template instead of rebuilding a workbook.")

uploaded = st.file_uploader("Upload Temu template Excel (.xlsx)", type=["xlsx"])

chunk_size = st.number_input("Data rows per file", min_value=100, max_value=999, value=999, step=1)
header_rows = st.number_input("Header rows to repeat", min_value=1, max_value=10, value=2, step=1)
sheet_name = st.text_input("Sheet name (optional)", value="")

if uploaded:
    st.success(f"Loaded: {uploaded.name}")

    progress_bar = st.progress(0)
    status = st.empty()

    def progress_callback(overall, part_num, total_parts, row_in_part, rows_in_part, elapsed):
        progress_bar.progress(min(max(overall, 0.0), 1.0))
        status.write(
            f"Progress: **{int(overall*100)}%** | File **{part_num}/{total_parts}** "
            f"| Row **{row_in_part}/{rows_in_part}** | Elapsed: **{int(elapsed)}s**"
        )

    if st.button("Split (Keep Format)"):
        try:
            zip_bytes, parts = split_using_template_keep_format(
                template_bytes=uploaded.read(),
                chunk_size=int(chunk_size),
                header_rows=int(header_rows),
                sheet_name=sheet_name.strip() or None,
                progress_callback=progress_callback,
            )

            progress_bar.progress(1.0)
            status.write(f"✅ Done. Created **{parts}** files with original formatting preserved.")

            st.download_button(
                "Download ZIP",
                data=zip_bytes,
                file_name="temu_split_keep_format.zip",
                mime="application/zip",
            )
        except Exception as e:
            st.error(f"Error: {e}")
